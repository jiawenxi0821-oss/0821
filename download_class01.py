import requests
import os
from openpyxl import load_workbook

# Excel 文件名
excel_file = "计算机视觉基础-班级01-自拍签到即为同意用于课程设计（不计入签到次数）.xlsx"

# 输出文件夹
output_dir = "签到照片"
os.makedirs(output_dir, exist_ok=True)

print("正在读取Excel文件并提取超链接...")
print(f"数据源: {excel_file}")

# 使用openpyxl加载工作簿以获取超链接
wb = load_workbook(excel_file)
ws = wb.active

# 获取标题行
headers_row = [cell.value for cell in ws[1]]

# 找到列索引
name_col_idx = None
id_col_idx = None
url_col_idx = None

for i, header in enumerate(headers_row):
    if header == "姓名":
        name_col_idx = i
    elif header == "学号/工号":
        id_col_idx = i
    elif header == "详情":
        url_col_idx = i

if name_col_idx is None or id_col_idx is None or url_col_idx is None:
    print("错误：找不到必需的列（姓名、学号/工号或详情）")
    exit(1)

print(f"找到列索引 - 姓名: {name_col_idx}, 学号/工号: {id_col_idx}, 详情: {url_col_idx}")

# 设置请求头，模拟浏览器访问
req_headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "image/webp,image/apng,image/*,*/*;q=0.8",
    "Referer": "https://p.cldisk.com/",
}

# 创建会话对象
session = requests.Session()
session.headers.update(req_headers)

# 从第二行开始处理数据
print("正在下载照片…")
success_count = 0
total_count = 0

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    total_count += 1

    name = str(row[name_col_idx].value).strip() if row[name_col_idx].value else ""
    sid = str(row[id_col_idx].value).strip() if row[id_col_idx].value else ""

    url_cell = row[url_col_idx]
    url = ""
    if url_cell.hyperlink:
        url = url_cell.hyperlink.target
    elif url_cell.value and str(url_cell.value).startswith("http"):
        url = str(url_cell.value)

    if not name or not sid:
        continue

    if not url or not url.startswith("http"):
        print(f"跳过无效 URL - {name} ({sid})")
        continue

    filename = f"{sid}-{name}.png"
    save_path = os.path.join(output_dir, filename)

    if os.path.exists(save_path):
        print(f"文件已存在，跳过: {filename}")
        success_count += 1
        continue

    try:
        print(f"正在下载: {name} ({sid})")
        r = session.get(url, timeout=15)
        
        if r.status_code == 200:
            with open(save_path, "wb") as f:
                f.write(r.content)
            print(f"成功：{filename}")
            success_count += 1
        else:
            print(f"失败（状态码 {r.status_code}）：{filename}")
    except Exception as e:
        print(f"下载错误 {filename}: {e}")

print(f"\n全部完成！成功下载: {success_count}/{total_count} 张照片")
print(f"照片已保存到：{output_dir}")
