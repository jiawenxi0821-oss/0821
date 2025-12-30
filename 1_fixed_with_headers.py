import pandas as pd
import requests
import os
from urllib.parse import urlparse
from openpyxl import load_workbook

# Excel 文件名
excel_file = "计算机视觉基础-班级02-自拍签到即为同意用于课程设计（不计入签到次数）.xlsx"

# 输出文件夹
output_dir = "签到照片2"
os.makedirs(output_dir, exist_ok=True)

print("正在读取Excel文件并提取超链接...")

# 使用openpyxl加载工作簿以获取超链接
wb = load_workbook(excel_file)
ws = wb.active

# 获取标题行
headers = [cell.value for cell in ws[1]]

# 找到列索引
name_col_idx = None
id_col_idx = None
url_col_idx = None

for i, header in enumerate(headers):
    if header == "姓名":
        name_col_idx = i
    elif header == "学号/工号":
        id_col_idx = i
    elif header == "详情":
        url_col_idx = i

if name_col_idx is None or id_col_idx is None or url_col_idx is None:
    print("错误：找不到必需的列（姓名、学号/工号或详情）")
    exit(1)

print(f"找到列索引 - 姓名: {name_col_idx}, 学号/工号: {id_col_idx}, 详情: {url_col_idx}")

# 设置请求头，模拟浏览器访问
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Accept": "image/webp,image/apng,image/*,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Referer": "https://p.cldisk.com/",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

# 创建会话对象，保持连接
session = requests.Session()
session.headers.update(headers)

# 从第二行开始处理数据
print("正在下载照片…")
success_count = 0
total_count = 0

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    total_count += 1

    # 获取单元格值
    name = str(row[name_col_idx].value).strip() if row[name_col_idx].value else ""
    sid = str(row[id_col_idx].value).strip() if row[id_col_idx].value else ""

    # 获取超链接
    url_cell = row[url_col_idx]
    url = ""
    if url_cell.hyperlink:
        url = url_cell.hyperlink.target
    elif url_cell.value and str(url_cell.value).startswith("http"):
        url = str(url_cell.value)

    if not name or not sid:
        print(f"跳过空姓名或学号: 行 {row_idx}")
        continue

    if not url or not url.startswith("http"):
        print(f"跳过无效 URL：{url if url else '(空)'} - {name} ({sid})")
        continue

    # 文件名：学号-姓名.png
    filename = f"{sid}-{name}.png"
    save_path = os.path.join(output_dir, filename)

    # 如果文件已存在，跳过
    if os.path.exists(save_path):
        print(f"文件已存在，跳过: {filename}")
        success_count += 1
        continue

    try:
        print(f"正在下载: {name} ({sid}) - {url}")
        
        # 增加重试机制
        max_retries = 3
        retry_count = 0
        success = False
        
        while retry_count < max_retries and not success:
            try:
                r = session.get(url, timeout=15)  # 增加超时时间
                
                # 如果403错误，尝试添加更多请求头
                if r.status_code == 403:
                    print("遇到403错误，尝试添加更多请求头...")
                    additional_headers = {
                        "Sec-Fetch-Dest": "image",
                        "Sec-Fetch-Mode": "no-cors",
                        "Sec-Fetch-Site": "same-origin",
                        "Cache-Control": "max-age=0"
                    }
                    session.headers.update(additional_headers)
                    r = session.get(url, timeout=15)
                
                if r.status_code == 200:
                    with open(save_path, "wb") as f:
                        f.write(r.content)
                    print(f"成功：{filename}")
                    success_count += 1
                    success = True
                else:
                    print(f"失败（状态码 {r.status_code}）：{filename} (尝试 {retry_count+1}/{max_retries})")
                    # 尝试获取更多信息
                    try:
                        error_info = r.text[:200]  # 只取前200个字符
                        print(f"错误信息: {error_info}")
                    except:
                        pass
                    retry_count += 1
                    
            except requests.exceptions.RequestException as req_e:
                print(f"请求错误 {filename} (尝试 {retry_count+1}/{max_retries}): {req_e}")
                retry_count += 1
                # 添加延迟避免频繁请求
                import time
                time.sleep(1)
                
        if not success:
            print(f"最终失败：{filename} (已尝试 {max_retries} 次)")
            
    except Exception as e:
        print(f"下载错误 {filename}: {e}")
        # 继续处理下一张照片，不中断整个程序

print(f"\n全部完成！成功下载: {success_count}/{total_count} 张照片")
print(f"照片已保存到：{output_dir}")

if success_count == 0:
    print("\n提示：所有图片下载失败，可能的原因：")
    print("1. 图片服务器需要认证或登录")
    print("2. 图片链接已过期")
    print("3. 图片服务器有严格的防爬虫机制")
    print("\n建议：")
    print("1. 尝试手动下载几张图片，确认链接是否有效")
    print("2. 如果手动下载成功，可能需要更复杂的请求头或认证机制")
